from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook

def get_excel_path():
    """Getting the file (GUI)"""
    Tk().withdraw()
    return askopenfilename(filetypes=[('Excel Files', '*.csv;*.xlsx;*.xlsm;*.xltx;*.xltm')],
                           title='Select your Excel File and Click Open')


path = get_excel_path()
wb = load_workbook(path)
ws = wb.active
unique_id = []

# crawling through all rows, and if the keyword matches then it will send the email. One email per email-id
# irrespective of number of occurrences
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    print(row)
    print(row.index('email'))