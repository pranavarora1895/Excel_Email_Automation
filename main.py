"""
EXCEL AND EMAIL AUTOMATION

Step 1: Goto myaccount.google.com/lesssecureapps and toggle the Less Secure Apps to allow python to send the
emails for you.
Step 2: Setup your email address and password as your environment variables to maintain your privacy.

"""
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
from openpyxl import load_workbook
import os
import smtplib

# setting email and password as env variables for privacy purposes
from openpyxl.utils.exceptions import InvalidFileException

EMAIL_ADDRESS = os.environ.get('EMAIL_ADDRESS')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASS')
try:
    def get_excel_path():
        """Getting the file (GUI)"""
        Tk().withdraw()
        return askopenfilename(filetypes=[('Excel Files', '*.csv;*.xlsx;*.xlsm;*.xltx;*.xltm')],
                               title='Select your Excel File and Click Open')
    def get_email_body_path():
        """Getting the file (GUI)"""
        Tk().withdraw()
        return askopenfilename(filetypes=[('All Files','*.*')],
                               title='Select your E-mail Body File and Click Open')

    path = get_excel_path()
    email_body = get_email_body_path()
    print('Processing......')
    def mail_sending_to_approved(receiver_email_id):
        """Function to send emails"""

        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

            subject = 'Congrats!! You got approved!'
            body = open(email_body, 'r').read()

            msg = f'Subject: {subject}\n\n{body}'

            smtp.sendmail(EMAIL_ADDRESS, receiver_email_id, msg)



    wb = load_workbook(path)
    ws = wb.active
    unique_id = []

    # getting the index number of email.
    email_idx = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        email_idx = row.index('email')

    # crawling through all rows, and if the keyword matches then it will send the email. One email per email-id
    # irrespective of number of occurrences
    keyword = input('Enter your keyword: ')
    print('Matching the keywords.......')
    print('Sending them mail......')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if keyword in row:
            if row[email_idx] in unique_id:
                continue
            else:
                mail_sending_to_approved(row[email_idx])
                unique_id.append(row[email_idx])
    if not unique_id:
        messagebox.showinfo("Email not Sent!!", "Your keyword didn't match with any id.")
    else:
        messagebox.showinfo("Email Sent!!", f'Email sent to {unique_id}')
except FileNotFoundError as e:
    print(f'Error: {e.errno}\nFile Not Found')
except OSError as o:
    print(f'Error: {o.errno}\nInvalid Entry')
except ValueError as v:
    print('ValueError')
except InvalidFileException as e:
    print('Cancelled')